import csv
from abc import ABC, abstractmethod
from typing import Generator

import openpyxl

from models.uploaded_file import UploadFileType


class FileReader(ABC):
    @abstractmethod
    def read_rows(self, file_path: str, file_type: UploadFileType) -> Generator[dict, None, None]:
        pass


class LocalFileReader(FileReader):
    def read_rows(self, file_path: str, file_type: UploadFileType) -> Generator[dict, None, None]:
        if file_type == UploadFileType.CSV:
            yield from self._read_csv(file_path)
        elif file_type == UploadFileType.EXCEL:
            yield from self._read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def _read_csv(self, file_path: str) -> Generator[dict, None, None]:
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield dict(row)

    def _read_excel(self, file_path: str) -> Generator[dict, None, None]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(headers, row))
        wb.close()


class CloudFileReader(FileReader):
    """
    Stub for future cloud storage support (S3, Azure Blob, GCS).
    Implement read_rows() by downloading the file to a temp path
    and delegating to LocalFileReader, or streaming directly.
    """
    def read_rows(self, file_path: str, file_type: UploadFileType) -> Generator[dict, None, None]:
        raise NotImplementedError("Cloud file reading is not yet implemented")


def get_file_reader(storage_backend: str = "local") -> FileReader:
    readers = {
        "local": LocalFileReader,
        "cloud": CloudFileReader,
    }
    reader_class = readers.get(storage_backend)
    if not reader_class:
        raise ValueError(f"Unknown storage backend: '{storage_backend}'. Choose from: {list(readers.keys())}")
    return reader_class()